import os
import io
import oss2
import pillow_avif # Registers AVIF support
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

def run_image_insertion(excel_file_stream, bucket_name, auth, endpoint):
    # 1. Setup OSS
    bucket = oss2.Bucket(auth, endpoint, bucket_name)
    
    # 2. Load Excel from the stream (from the frontend)
    wb = load_workbook(excel_file_stream)
    ws = wb.active

    # Insert column if missing
    if ws.cell(row=1, column=2).value != "Product Image":
        ws.insert_cols(2)
        ws.cell(row=1, column=2).value = "Product Image"

    temp_files = []
    extensions = ['.AVIF', '.avif', '.webp', '.WEBP', '.png', '.PNG', '.jpg', '.JPG', '.jpeg', '.JPEG']

    # 3. Loop through rows
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row, column=1).value).strip()
        if not sku or sku == "None": continue

        found_stream = None
        for ext in extensions:
            object_key = f"{sku}{ext}"
            if bucket.object_exists(object_key):
                found_stream = bucket.get_object(object_key).read()
                break

        if found_stream:
            try:
                # Check if it's an XML error instead of an image
                if found_stream.startswith(b'<?xml') or found_stream.startswith(b'<Error'):
                    continue

                with PILImage.open(io.BytesIO(found_stream)) as img:
                    rgb_img = img.convert('RGB')
                    temp_name = f"temp_{sku}.png"
                    rgb_img.save(temp_name)
                    temp_files.append(temp_name)

                    xl_img = ExcelImage(temp_name)
                    xl_img.width, xl_img.height = (100, 100)
                    ws.row_dimensions[row].height = 80
                    ws.add_image(xl_img, f"B{row}")
            except:
                continue 

    # 4. Save result to a memory buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    # 5. Cleanup temp files
    for f in temp_files:
        if os.path.exists(f): os.remove(f)
        
    return output_buffer